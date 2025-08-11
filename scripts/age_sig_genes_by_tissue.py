#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Build one Excel workbook with one sheet per tissue (only sheets that actually
have significant rows). Rows are filtered by BH_adjusted_age < 0.05.
Only rows with non-missing BH_adjusted_age are kept; other columns may contain NA.

Stability/performance:
- Write to a local temp file with openpyxl (fast, no cloud-sync conflicts),
  then atomically move to FINAL_XLSX in your project directory.
- Keep values as NUMERIC in Excel.
- After writing, apply per-cell number_format:
    * scientific notation if 0 < |value| < 0.1
    * normal decimal otherwise (including 0)
- Auto-adjust column widths and row heights for readability.

Summary sheet:
- Adds Up_count / Down_count / Up_pct / Down_pct per tissue, based on age_coef:
  >0 → Up, <0 → Down, ==0 or NaN → ignored for proportions.
"""

import pandas as pd
import glob, os, re, shutil, tempfile
from typing import List
from openpyxl.utils import get_column_letter

INPUT_DIR   = "/Users/chenmenghui/Desktop/shinyProject/datasets/pval_results/"
FINAL_XLSX  = os.path.join(INPUT_DIR, "age_sig_genes_by_tissue.xlsx")

PVAL_COL      = "BH_adjusted_age"
ALPHA         = 0.05          
WRITE_SUMMARY = True           

NUM_COLS: List[str] = [
    "intercept",
    "p_value_age",
    "BH_adjusted_age",
    "age_coef",
    "p_value_sex",
    "BH_adjusted_sex",
    "sex_coef",
]

SCI_THRESHOLD = 0.1
SCI_FMT = "0.00E+00"          
DEC_FMT = "0.###############" 

def find_pval_col(cols, target="BH_adjusted_age"):
    """Find the BH_adjusted_age column (exact, case-insensitive) or a close variant."""
    for c in cols:
        if c.strip().lower() == target.lower():
            return c
    cand = [c for c in cols
            if ("bh" in c.lower() and "adjust" in c.lower() and "age" in c.lower())]
    return cand[0] if cand else None

def clean_sheet_name(name, used):
    """Sanitize Excel sheet name and ensure uniqueness."""
    for suf in ["_pvalue_results", "-pvalue-results"]:
        if name.endswith(suf):
            name = name[: -len(suf)]
    name = re.sub(r'[\[\]\*\?\\/:\']', '_', name).strip()
    if len(name) > 31:
        name = name[:31]
    base = name
    i = 1
    while name in used:
        suffix = f"_{i}"
        allowed = 31 - len(suffix)
        name = (base[:allowed] + suffix) if len(base) + len(suffix) > 31 else base + suffix
        i += 1
    used.add(name)
    return name

def coerce_numeric_cols(df, cols):
    for col in cols:
        if col in df.columns:
            s = df[col].astype(str)
            s = s.str.replace(",", "", regex=False)   
            s = s.str.replace("%", "", regex=False)   
            s = s.str.strip()
            df[col] = pd.to_numeric(s, errors="coerce")
    return df

def format_sheet_numeric_cells_openpyxl(writer, sheet_name, df, numeric_cols,
                                        sci_threshold=SCI_THRESHOLD,
                                        sci_fmt=SCI_FMT, dec_fmt=DEC_FMT):

    ws = writer.book[sheet_name]
    col_to_idx = {col: i+1 for i, col in enumerate(df.columns)}  # 1-based

    for col in numeric_cols:
        if col not in df.columns:
            continue
        if not pd.api.types.is_numeric_dtype(df[col]):
            continue

        j = col_to_idx[col] 
        for i, val in enumerate(df[col].values, start=2):  
            if pd.isna(val):
                continue
            try:
                use_sci = (abs(val) < sci_threshold) and (val != 0)
                ws.cell(row=i, column=j).number_format = sci_fmt if use_sci else dec_fmt
            except Exception:
                pass

    for col_idx, col_name in enumerate(df.columns, start=1):
        values_as_str = [str(col_name)]
        col_series = df[col_name]
        if pd.api.types.is_numeric_dtype(col_series) and col_name in numeric_cols:
            def _disp_str(x):
                if pd.isna(x):
                    return ""
                return f"{x:.2e}" if (abs(x) < sci_threshold and x != 0) else f"{x}"
            values_as_str += [_disp_str(v) for v in col_series.values]
        else:
            values_as_str += [("" if pd.isna(v) else str(v)) for v in col_series.values]

        max_len = max((len(v) for v in values_as_str), default=0)
        adjusted_width = min(max(10, max_len + 2), 80)
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

    n_rows = df.shape[0]
    ws.row_dimensions[1].height = 20
    for r in range(2, n_rows + 2):
        ws.row_dimensions[r].height = 18

def main():
    csv_files = sorted(glob.glob(os.path.join(INPUT_DIR, "*.csv")))
    if not csv_files:
        print("No CSV files were found. Please check INPUT_DIR.")
        return

    sheet_names_used = set()
    summary_rows = []
    sheets_written = 0

    tmp_fd, tmp_path = tempfile.mkstemp(suffix=".xlsx")
    os.close(tmp_fd)

    try:
        with pd.ExcelWriter(tmp_path, engine="openpyxl") as writer:
            for f in csv_files:
                try:
                    df = pd.read_csv(f)
                except Exception as e:
                    summary_rows.append({
                        "File": os.path.basename(f),
                        "Status": f"read failed: {e}",
                        "Rows_total": None,
                        "Rows_kept": None,
                        "Up_count": None,
                        "Down_count": None,
                        "Up_pct": None,
                        "Down_pct": None,
                        "Sheet": None
                    })
                    continue

                df.columns = [c.strip() for c in df.columns]
                df = coerce_numeric_cols(df, NUM_COLS)

                # Locate BH_adjusted_age column
                pcol = find_pval_col(df.columns, PVAL_COL)
                if pcol is None:
                    summary_rows.append({
                        "File": os.path.basename(f),
                        "Status": "missing BH_adjusted_age column",
                        "Rows_total": int(df.shape[0]),
                        "Rows_kept": 0,
                        "Up_count": 0,
                        "Down_count": 0,
                        "Up_pct": None,
                        "Down_pct": None,
                        "Sheet": None
                    })
                    continue

                mask_valid = df[pcol].notna()
                mask_sig   = df[pcol] < ALPHA
                filtered   = df.loc[mask_valid & mask_sig].copy()

                kept  = int(filtered.shape[0])
                total = int(df.shape[0])

                up_cnt = down_cnt = 0
                up_pct = down_pct = None
                if kept > 0:
                    if "age_coef" in filtered.columns:
                        up_cnt = int((filtered["age_coef"] > 0).sum())
                        down_cnt = int((filtered["age_coef"] < 0).sum())
                        denom = up_cnt + down_cnt
                        if denom > 0:
                            up_pct = round(up_cnt / denom * 100, 2)
                            down_pct = round(down_cnt / denom * 100, 2)

                if kept > 0:
                    base  = os.path.splitext(os.path.basename(f))[0]
                    sheet = clean_sheet_name(base, sheet_names_used)
                    filtered.to_excel(writer, index=False, sheet_name=sheet)
                    format_sheet_numeric_cells_openpyxl(
                        writer, sheet, filtered, NUM_COLS,
                        sci_threshold=SCI_THRESHOLD,
                        sci_fmt=SCI_FMT,
                        dec_fmt=DEC_FMT
                    )

                    sheets_written += 1
                    summary_rows.append({
                        "File": os.path.basename(f),
                        "Status": "OK",
                        "Rows_total": total,
                        "Rows_kept": kept,
                        "Up_count": up_cnt,
                        "Down_count": down_cnt,
                        "Up_pct": up_pct,
                        "Down_pct": down_pct,
                        "Sheet": sheet
                    })
                else:
                    summary_rows.append({
                        "File": os.path.basename(f),
                        "Status": "no age-significant rows (no sheet created)",
                        "Rows_total": total,
                        "Rows_kept": kept,
                        "Up_count": 0,
                        "Down_count": 0,
                        "Up_pct": None,
                        "Down_pct": None,
                        "Sheet": None
                    })

            if WRITE_SUMMARY:
                summary_df = pd.DataFrame(summary_rows)
                cols = ["File","Status","Rows_total","Rows_kept",
                        "Up_count","Down_count","Up_pct","Down_pct","Sheet"]
                summary_df = summary_df[cols]
                summary_df.to_excel(writer, index=False, sheet_name="Summary")

        shutil.move(tmp_path, FINAL_XLSX)
        print(f"Completed: wrote {sheets_written} sheets with data → {FINAL_XLSX}")

    finally:
        try:
            if os.path.exists(tmp_path):
                os.remove(tmp_path)
        except Exception:
            pass

if __name__ == "__main__":
    main()